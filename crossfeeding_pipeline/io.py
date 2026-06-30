"""
io.py
=====
Excel sheet loader. Thin wrapper around openpyxl that pulls:

    * feature names              (row labelled "Sample")
    * CASIC metabolite names     (row labelled "CASIC_Result", fallback row 5)
    * metabolite class strings   (row labelled "Metabolite_Class", fallback row 4)
    * all sample rows whose name begins with one of the valid media prefixes
      (QC / NIST_Name / Input_Metabolite etc. are silently filtered out)

No normalisation, no mean/CV. That work is done in preprocessing.py so
each module has one responsibility.
"""

import numpy as np
import pandas as pd
import openpyxl


def _safe_float(value):
    """Convert a cell value to float, returning NaN on failure."""
    try:
        return float(value)
    except (TypeError, ValueError):
        return np.nan


def load_sheet(xlsx_path, sheet_name, media_prefixes):
    """
    Load one medium sheet from the Excel workbook.

    Parameters
    ----------
    xlsx_path : str or Path
        Path to the Excel workbook.
    sheet_name : str
        Sheet (medium) to load — one of AUM / DEX / GLY / HisGly / SUC.
    media_prefixes : list[str]
        Valid medium prefixes. A sample row is kept only if its row-label
        starts with "{prefix}_" for at least one prefix. This matches the
        original v2.0 bug fix #1 (keeps real samples, drops metadata rows).

    Returns
    -------
    raw_df     : DataFrame (sample x feature, technical replicates kept)
    casic_map  : dict {feature_name: CASIC_Result string}
    class_map  : dict {feature_name: metabolite class string}
    bio_feats  : list[str] — biological feature columns (everything not "(IS)")
    is_feats   : list[str] — internal-standard feature columns
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]

    # ── Index first-column labels (row-headers) so we can find the layout ──
    label_to_row = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        key = str(v).strip()
        if key and key not in label_to_row:
            label_to_row[key] = r

    if "Sample" not in label_to_row:
        raise ValueError(
            f"[{sheet_name}] Sheet layout error: row 'Sample' not found. "
            f"Found row labels: {list(label_to_row.keys())[:20]}..."
        )

    feat_row  = label_to_row["Sample"]
    casic_row = label_to_row.get("CASIC_Result",     5)
    class_row = label_to_row.get("Metabolite_Class", 4)
    n_cols    = ws.max_column

    # ── Column headers ───────────────────────────────────────────────────────
    feat_names = [str(ws.cell(feat_row,  c).value or "").strip()
                  for c in range(2, n_cols + 1)]
    casic_vals = [str(ws.cell(casic_row, c).value or "").strip()
                  for c in range(2, n_cols + 1)]
    class_vals = [str(ws.cell(class_row, c).value or "Unknown").strip()
                  for c in range(2, n_cols + 1)]

    casic_map = dict(zip(feat_names, casic_vals))
    class_map = dict(zip(feat_names, class_vals))

    is_feats  = [f for f in feat_names if f.startswith("(IS)")]
    bio_feats = [f for f in feat_names if f and not f.startswith("(IS)")]

    # ── Sample rows ──────────────────────────────────────────────────────────
    valid_prefixes = tuple(f"{m}_" for m in media_prefixes)
    sample_rows = []
    for r in range(feat_row + 1, ws.max_row + 1):
        name = str(ws.cell(r, 1).value or "").strip()
        if not name or name == "nan" or name.startswith("QC"):
            continue
        if not name.startswith(valid_prefixes):
            continue
        vals = [_safe_float(ws.cell(r, c).value) for c in range(2, n_cols + 1)]
        sample_rows.append((name, vals))

    raw_df = pd.DataFrame(
        [v for _, v in sample_rows],
        index=[s for s, _ in sample_rows],
        columns=feat_names,
    )
    # Drop any columns with blank names (trailing Excel empties)
    raw_df = raw_df.loc[:, raw_df.columns != ""]
    bio_feats = [f for f in bio_feats if f != ""]
    is_feats = [f for f in is_feats if f != ""]

    wb.close()
    return raw_df, casic_map, class_map, bio_feats, is_feats
