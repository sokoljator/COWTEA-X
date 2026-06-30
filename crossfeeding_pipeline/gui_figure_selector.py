"""
gui_figure_selector.py
======================
Lightweight Tkinter GUI for browsing pipeline results and producing
publication-quality figures.

Workflow:
  1. Browse to a results folder (one of the timestamped output dirs).
  2. The GUI loads `All_ScoredCandidates_enhanced.csv` (preferred) or
     `All_ScoredCandidates.csv` if the enhanced file isn't there.
  3. Filter by score threshold, medium, chain, pattern label.
  4. Pick a metabolite + chain pair, choose the plot method, and click
     "Render" to save a high-res figure into a `figures_gui/` subfolder.

Plot methods supported:
  * Abundance bars (uses the fixed `viz_abundance.fig_abundance_bars`)
  * Reuse one of the per-medium figures already saved in the folder
    (the GUI just opens the existing PNG)

Running:
    python -m crossfeeding_pipeline.gui_figure_selector

The module imports cleanly without a display attached — Tk is only
created when `main()` runs.
"""

from __future__ import annotations
from pathlib import Path
from typing import Optional
import sys

import numpy as np
import pandas as pd


# Allow both supported launch styles:
#   1) python -m crossfeeding_pipeline.gui_figure_selector
#   2) python crossfeeding_pipeline/gui_figure_selector.py
#
# In style #2, Python sets __package__ to None and does not put the package's
# parent directory on sys.path, so absolute imports such as
# `crossfeeding_pipeline.viz_abundance` fail.  This small bootstrap makes the
# direct-file launch used by PyCharm work without changing normal package use.
if __package__ in (None, ""):
    _PKG_DIR = Path(__file__).resolve().parent
    _PARENT = _PKG_DIR.parent
    if str(_PARENT) not in sys.path:
        sys.path.insert(0, str(_PARENT))


def _load_results(folder: Path) -> tuple[pd.DataFrame, str]:
    """Return (df, which_file)."""
    cand_files = ["All_ScoredCandidates_enhanced.csv",
                  "All_ScoredCandidates.csv"]
    for fn in cand_files:
        p = folder / fn
        if p.exists():
            df = pd.read_csv(p)
            return _augment_metabolite_metadata(df, folder), fn
    raise FileNotFoundError(
        f"Neither {cand_files[0]} nor {cand_files[1]} found in {folder}"
    )


def _augment_metabolite_metadata(df: pd.DataFrame, folder: Path) -> pd.DataFrame:
    """
    Add CASIC/user-facing metabolite labels if the result CSV does not already
    contain them. Newer pipeline runs write `metabolite_name` and
    `metabolite_class` directly. Older enhanced runs did not, so this function
    reconstructs the mapping from the original workbook path stored in
    debug_report.json.
    """
    if df.empty or "metabolite" not in df.columns:
        return df
    if "metabolite_name" in df.columns and df["metabolite_name"].notna().any():
        return df

    import json
    try:
        import openpyxl
    except Exception:
        return df

    debug_path = Path(folder) / "debug_report.json"
    if not debug_path.exists():
        return df

    try:
        debug = json.loads(debug_path.read_text(encoding="utf-8"))
        xlsx_path = Path(debug.get("config_snapshot", {}).get("xlsx_path", ""))
    except Exception:
        return df
    if not xlsx_path.exists():
        return df

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception:
        return df

    name_map = {}
    class_map = {}
    for med in sorted(df.get("medium", pd.Series(dtype=str)).dropna().astype(str).unique()):
        if med not in wb.sheetnames:
            continue
        ws = wb[med]
        row_labels = {str(ws.cell(row=r, column=1).value): r
                      for r in range(1, min(ws.max_row, 15) + 1)}
        casic_r = row_labels.get("CASIC_Result")
        class_r = row_labels.get("Metabolite_Class")
        input_r = row_labels.get("Input_Metabolite")
        if not casic_r:
            continue
        for c in range(2, ws.max_column + 1):
            header_name = ws.cell(row=1, column=c).value
            input_name = ws.cell(row=input_r, column=c).value if input_r else None
            casic_name = ws.cell(row=casic_r, column=c).value
            class_name = ws.cell(row=class_r, column=c).value if class_r else None
            for raw in (header_name, input_name):
                if raw is None:
                    continue
                raw = str(raw)
                if casic_name is not None and str(casic_name).strip():
                    name_map[raw] = str(casic_name)
                if class_name is not None and str(class_name).strip():
                    class_map[raw] = str(class_name)

    if name_map:
        df = df.copy()
        df["metabolite_name"] = df["metabolite"].map(name_map).fillna(df["metabolite"])
        if "metabolite_class" not in df.columns:
            df["metabolite_class"] = df["metabolite"].map(class_map)
    return df


def _display_metabolite(row: pd.Series) -> str:
    """Return the user-facing metabolite name, preferably CASIC_Result."""
    for col in ("metabolite_name", "CASIC_Result", "casic_result", "casic_name"):
        if col in row.index and pd.notna(row.get(col)) and str(row.get(col)).strip():
            return str(row.get(col))
    return str(row.get("metabolite", ""))


def _safe_name(name: str) -> str:
    """Match the filename sanitisation used for abundance-bar PNGs."""
    import re
    return re.sub(r"[^\w]", "_", str(name))[:40]


def _safe_filter(df: pd.DataFrame, col: str, value: str) -> pd.DataFrame:
    if value in ("(any)", "", None):
        return df
    if col not in df.columns:
        return df
    return df[df[col].astype(str) == value]


def _selectable_values(df: pd.DataFrame, col: str) -> list[str]:
    if col not in df.columns:
        return ["(any)"]
    vals = sorted(df[col].dropna().astype(str).unique())
    return ["(any)"] + vals


def main():
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    root = tk.Tk()
    root.title("COWTEA-X Figure Selector")
    root.geometry("780x620")

    state = {"folder": None, "df": pd.DataFrame(), "shown_rows": {}}

    # ── Top: folder picker ──
    top = ttk.Frame(root, padding=8)
    top.pack(fill="x")
    ttk.Label(top, text="Results folder:").pack(side="left")
    folder_var = tk.StringVar()
    ttk.Entry(top, textvariable=folder_var, width=70).pack(side="left", padx=4)

    def pick_folder():
        f = filedialog.askdirectory(title="Pick a COWTEA-X output folder")
        if not f:
            return
        folder_var.set(f)
        try:
            df, used = _load_results(Path(f))
        except Exception as e:
            messagebox.showerror("Load error", str(e))
            return
        state["folder"] = Path(f)
        state["df"] = df
        status_var.set(f"Loaded {used}: {len(df):,} rows")
        _refresh_filters()
        _apply_filter()

    ttk.Button(top, text="Browse...", command=pick_folder).pack(side="left")

    # ── Filters ──
    filt = ttk.LabelFrame(root, text="Filters", padding=8)
    filt.pack(fill="x", padx=8, pady=4)

    score_var = tk.DoubleVar(value=3.0)
    ttk.Label(filt, text="min score:").grid(row=0, column=0, sticky="w")
    ttk.Spinbox(filt, from_=0, to=12, increment=1.0, textvariable=score_var,
                width=6).grid(row=0, column=1, padx=4)

    medium_var  = tk.StringVar(value="(any)")
    chain_var   = tk.StringVar(value="(any)")
    pattern_var = tk.StringVar(value="(any)")

    ttk.Label(filt, text="medium:").grid(row=0, column=2)
    medium_cb = ttk.Combobox(filt, textvariable=medium_var,
                             width=10, state="readonly")
    medium_cb.grid(row=0, column=3, padx=4)

    ttk.Label(filt, text="chain:").grid(row=0, column=4)
    chain_cb  = ttk.Combobox(filt, textvariable=chain_var,
                             width=12, state="readonly")
    chain_cb.grid(row=0, column=5, padx=4)

    ttk.Label(filt, text="pattern:").grid(row=0, column=6)
    pattern_cb = ttk.Combobox(filt, textvariable=pattern_var,
                              width=12, state="readonly")
    pattern_cb.grid(row=0, column=7, padx=4)

    def _refresh_filters():
        df = state["df"]
        medium_cb["values"]  = _selectable_values(df, "medium")
        chain_cb["values"]   = _selectable_values(df, "chain")
        pattern_cb["values"] = _selectable_values(df, "pattern_label")

    # ── Results table ──
    tbl_frame = ttk.LabelFrame(root, text="Candidate rows", padding=4)
    tbl_frame.pack(fill="both", expand=True, padx=8, pady=4)

    cols = ("medium", "chain", "CASIC_name", "raw_metabolite", "score",
            "pattern_label", "confidence")
    tree = ttk.Treeview(tbl_frame, columns=cols, show="headings", height=14)
    for c in cols:
        tree.heading(c, text=c)
        tree.column(c, width=120, anchor="w")
    tree.pack(side="left", fill="both", expand=True)
    sb = ttk.Scrollbar(tbl_frame, orient="vertical", command=tree.yview)
    sb.pack(side="right", fill="y")
    tree.configure(yscrollcommand=sb.set)

    def _apply_filter(*_):
        df = state["df"]
        if df.empty:
            tree.delete(*tree.get_children())
            return
        f = df.copy()
        if "score" in f.columns:
            f = f[f["score"] >= float(score_var.get())]
        f = _safe_filter(f, "medium", medium_var.get())
        f = _safe_filter(f, "chain",  chain_var.get())
        f = _safe_filter(f, "pattern_label", pattern_var.get())
        f = f.head(500)  # cap UI rows
        tree.delete(*tree.get_children())
        state["shown_rows"] = {}
        for _, row in f.iterrows():
            display_name = _display_metabolite(row)
            iid = tree.insert("", "end", values=(
                row.get("medium", ""),
                row.get("chain", ""),
                display_name,
                row.get("metabolite", ""),
                row.get("score", ""),
                row.get("pattern_label", row.get("category", "")),
                row.get("confidence", ""),
            ))
            state["shown_rows"][iid] = row.to_dict() | {"_display_metabolite": display_name}
        status_var.set(f"Filter → {len(f):,} rows shown")

    for v in (score_var, medium_var, chain_var, pattern_var):
        v.trace_add("write", lambda *_: _apply_filter())

    # ── Render controls ──
    ctl = ttk.LabelFrame(root, text="Render selected", padding=8)
    ctl.pack(fill="x", padx=8, pady=4)

    method_var = tk.StringVar(value="abundance_bars")
    ttk.Label(ctl, text="method:").pack(side="left")
    ttk.Combobox(ctl, textvariable=method_var,
                 values=["abundance_bars", "open_existing_PNG"],
                 width=24, state="readonly").pack(side="left", padx=4)

    dpi_var = tk.IntVar(value=300)
    ttk.Label(ctl, text="DPI:").pack(side="left", padx=(12, 2))
    ttk.Spinbox(ctl, from_=150, to=600, increment=50, textvariable=dpi_var,
                width=5).pack(side="left")

    def render():
        sel = tree.selection()
        if not sel:
            messagebox.showinfo("No selection", "Pick a row first.")
            return
        row = state["shown_rows"].get(sel[0], {})
        vals = tree.item(sel[0], "values")
        med = str(row.get("medium", vals[0]))
        chain = str(row.get("chain", vals[1]))
        raw_metab = str(row.get("metabolite", vals[3]))
        display_metab = str(row.get("_display_metabolite", vals[2]))
        folder = state["folder"]
        if folder is None:
            return

        if method_var.get() == "abundance_bars":
            ok, msg = _render_abundance_bars(folder, med, chain, raw_metab,
                                             display_metab=display_metab,
                                             dpi=int(dpi_var.get()))
        else:
            ok, msg = _open_existing_png(folder, med, chain, display_metab)
        if ok:
            messagebox.showinfo("Render OK", msg)
        else:
            messagebox.showwarning("Render failed", msg)

    ttk.Button(ctl, text="Render", command=render).pack(side="right")

    # ── Status bar ──
    status_var = tk.StringVar(value="No folder loaded.")
    ttk.Label(root, textvariable=status_var, relief="sunken", anchor="w")\
        .pack(side="bottom", fill="x")

    root.mainloop()


# ── render helpers ──────────────────────────────────────────────────────
def _render_abundance_bars(folder: Path, med: str, chain: str, metab: str,
                           display_metab: Optional[str] = None,
                           dpi: int = 300) -> tuple[bool, str]:
    """
    Generate an abundance-bars PNG for one (med, chain, metabolite) by
    re-loading the per-medium IS-normalised stats. We rely on the
    `IS_normalized_*.xlsx` exports in the same folder for the means/SDs
    (those are produced by `export_excel`).
    """
    import re
    try:
        from .viz_abundance import fig_abundance_bars
    except ImportError:
        from crossfeeding_pipeline.viz_abundance import fig_abundance_bars

    out_dir = folder / "figures_gui"
    out_dir.mkdir(exist_ok=True)

    # Locate the IS-normalised export for this medium
    is_xlsx = next(folder.glob(f"IS_normalized_{med}*.xlsx"), None)
    class_map = {}
    if is_xlsx is not None:
        try:
            xls = pd.ExcelFile(is_xlsx)
            mean_df = pd.read_excel(xls, "mean",  index_col=0)
            sd_df   = pd.read_excel(xls, "sd",    index_col=0) if "sd" in xls.sheet_names \
                      else None
            cv_df   = pd.read_excel(xls, "cv",    index_col=0) if "cv" in xls.sheet_names \
                      else None
        except Exception as e:
            return False, f"Failed reading {is_xlsx.name}: {e}"
    else:
        try:
            mean_df, cv_df, sd_df, _, class_map = _load_stats_from_source(folder, med)
        except Exception as e:
            return False, (
                f"Could not find IS_normalized_{med}*.xlsx and could not "
                f"recompute stats from the original workbook: {e}"
            )

    # Derive the 7 sample keys from chain name
    chain_to_strain = {
        "CLN-513p": ("Tmp05", "Tme13"),
        "CLN-513e": ("Tme13", "Tmp05"),
        "CLN-614p": ("Tmp06", "Tme14"),
        "CLN-614e": ("Tme14", "Tmp06"),
        "CTR-1p":   ("PA01",  "K12"),
        "CTR-1e":   ("K12",   "PA01"),
    }
    if chain not in chain_to_strain:
        return False, f"Unknown chain ID: {chain}"

    first, second = chain_to_strain[chain]
    # When the chain is a P-chain, P=first; E=second. For E-chain, vice versa.
    if chain.endswith("p"):
        p_strain, e_strain = first, second
    else:
        e_strain, p_strain = first, second
    p_chain_name = next((c for c, (a, _) in chain_to_strain.items()
                         if c.endswith("p") and a == p_strain), None)
    e_chain_name = next((c for c, (a, _) in chain_to_strain.items()
                         if c.endswith("e") and a == e_strain), None)

    row_keys = [
        f"{med}_medium",
        f"{med}_3H_{p_strain}",
        f"{med}_5H_{p_strain}_{e_strain}",
        f"{med}_20H_{p_strain}_{e_strain}_{p_strain}",
        f"{med}_3H_{e_strain}",
        f"{med}_5H_{e_strain}_{p_strain}",
        f"{med}_20H_{e_strain}_{p_strain}_{e_strain}",
    ]
    missing = [k for k in row_keys if k not in mean_df.index]
    if missing:
        return False, f"Missing samples in IS_normalized_{med}: {missing}"
    if metab not in mean_df.columns:
        return False, f"Metabolite '{metab}' not found in mean sheet"

    m7 = mean_df.loc[row_keys, metab].values.astype(float)
    s7 = (sd_df.loc[row_keys, metab].values.astype(float)
          if sd_df is not None and metab in sd_df.columns else None)
    c7 = (cv_df.loc[row_keys, metab].values.astype(float)
          if cv_df is not None and metab in cv_df.columns else None)

    display_metab = display_metab or metab
    safe = _safe_name(display_metab)
    out_path = out_dir / f"GUI_{med}_{p_chain_name}_vs_{e_chain_name}_{safe}.png"
    metab_class = None
    try:
        metab_class = class_map.get(metab)
    except Exception:
        metab_class = None
    p = fig_abundance_bars(
        m7, s7, c7,
        chain_p_name=p_chain_name, chain_e_name=e_chain_name,
        metabolite=metab, casic_name=display_metab, metab_class=metab_class,
        condition=med, output_path=out_path, dpi=dpi,
    )
    if p is None:
        return False, "Plot was skipped (no valid signal)."
    return True, f"Saved: {p}"


def _open_existing_png(folder: Path, med: str, chain: str, metab: str
                       ) -> tuple[bool, str]:
    safe = _safe_name(metab)
    matches = list(folder.glob(f"AbundanceBars_{med}_*{chain}*{safe}*.png"))
    if not matches:
        # Fallback for older plots or Unicode/filename differences.
        chain_matches = list(folder.glob(f"AbundanceBars_{med}_*{chain}*.png"))
        matches = [p for p in chain_matches if safe.lower() in p.stem.lower()]
    if not matches:
        return False, f"No AbundanceBars_*{med}*{chain}*{safe}*.png in {folder}"
    import webbrowser
    webbrowser.open(str(matches[0]))
    return True, f"Opened: {matches[0].name}"


def _load_stats_from_source(folder: Path, med: str):
    """
    Recompute mean/CV/SD matrices from the source workbook referenced in
    debug_report.json. This is used by the GUI when the intermediate
    IS_normalized_*.xlsx file is not present in the result folder.
    """
    import json
    try:
        from .constants import MEDIA
        from .io import load_sheet
        from .preprocessing import normalize_is, compute_stats
    except ImportError:
        from crossfeeding_pipeline.constants import MEDIA
        from crossfeeding_pipeline.io import load_sheet
        from crossfeeding_pipeline.preprocessing import normalize_is, compute_stats

    debug_path = Path(folder) / "debug_report.json"
    if not debug_path.exists():
        raise FileNotFoundError("debug_report.json is missing")
    debug = json.loads(debug_path.read_text(encoding="utf-8"))
    cfg = debug.get("config_snapshot", {}) or {}
    xlsx_path = Path(cfg.get("xlsx_path", ""))
    if not xlsx_path.exists():
        raise FileNotFoundError(
            f"source workbook path from debug_report.json does not exist: {xlsx_path}"
        )
    media = cfg.get("media") or MEDIA
    raw_df, casic_map, class_map, bio_feats, is_feats = load_sheet(
        xlsx_path, med, media
    )
    norm_df, _stable_is = normalize_is(
        raw_df, is_feats, bio_feats, sheet_name=med, verbose=False
    )
    mean_df, cv_df, sd_df = compute_stats(norm_df)
    return mean_df, cv_df, sd_df, casic_map, class_map


if __name__ == "__main__":
    sys.exit(main())
