"""
export.py
=========
Per-medium Excel export — produces a 5-sheet workbook:

    1. log2FC_vs_blank         — wide FC matrix (sample x metabolite),
                                  CASIC names used as column headers
    2. CrossFeeding_Candidates — simple step-FC candidates from analysis.py
    3. ScoredCandidates        — full scoring output (pattern + PGM)
    4. CV_Table                — CV of 2 technical reps, cells with
                                  CV > CV_THRESH highlighted light red
    5. Summary                 — scored rows with score >= 2, coloured
                                  by score band for quick triage

Ported from crossfeeding_pipeline_v2.0.py :: export_excel (lines 3060-3205)
with only trivial renames: `log2fc` now comes from utils, `CV_THRESH` and
`MEDIA` from constants. No logic changes.
"""

from pathlib import Path

import numpy as np
import pandas as pd

from .constants import MEDIA, CV_THRESH
from .utils import log2fc


def export_excel(mean_df, cv_df, candidates_df, scored_df,
                 casic_map, class_map, bio_feats, medium, out_dir):
    """
    Write Results_{medium}.xlsx with 5 sheets (see module docstring).

    All arguments exactly match the v2.0 signature so downstream callers
    do not need to change.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    out = Path(out_dir) / f"Results_{medium}.xlsx"
    writer = pd.ExcelWriter(out, engine="openpyxl")
    blank_key = f"{medium}_medium"
    feats  = [f for f in bio_feats if f in mean_df.columns]

    # Metadata-row filter — only sample rows (prefix MED_) reach the sheets
    valid_prefixes = tuple(f"{m}_" for m in MEDIA)

    # ── Sheet 1: log2FC vs blank ────────────────────────────────────────────
    if blank_key in mean_df.index:
        blank = mean_df.loc[blank_key, feats]
        valid_sample_idx = [s for s in mean_df.index
                            if s != blank_key and str(s).startswith(valid_prefixes)]
        fc_rows = {s: log2fc(mean_df.loc[s, feats], blank).round(3)
                   for s in valid_sample_idx}
        fc_df = pd.DataFrame(fc_rows).T
        fc_df.index.name = "Sample"
        fc_df.columns = [casic_map.get(c, c) for c in fc_df.columns]
        fc_df.to_excel(writer, sheet_name="log2FC_vs_blank")

    # ── Sheet 2: original step-FC candidates ────────────────────────────────
    if candidates_df is not None and not candidates_df.empty:
        cand_out = candidates_df.copy()
        cand_out["metabolite_name"]  = cand_out["metabolite"].map(casic_map)
        cand_out["metabolite_class"] = cand_out["metabolite"].map(class_map)
        cand_out.to_excel(writer, sheet_name="CrossFeeding_Candidates", index=False)

    # ── Sheet 3: full scored candidates ─────────────────────────────────────
    if scored_df is not None and not scored_df.empty:
        sc_out = scored_df.copy()
        sc_out["metabolite_name"]  = sc_out["metabolite"].map(casic_map)
        sc_out["metabolite_class"] = sc_out["metabolite"].map(class_map)
        sc_out.to_excel(writer, sheet_name="ScoredCandidates", index=False)

    # ── Sheet 4: CV table (sample rows only) ────────────────────────────────
    valid_cv_idx = [s for s in cv_df.index if str(s).startswith(valid_prefixes)]
    cv_out = (cv_df.loc[valid_cv_idx, feats].copy()
              if valid_cv_idx else cv_df[feats].copy())
    cv_out.index.name = "Sample"
    cv_out.columns = [casic_map.get(c, c) for c in cv_out.columns]
    cv_out.to_excel(writer, sheet_name="CV_Table")

    # ── Sheet 5: Summary (score >= 2, with FC_produced / FC_consumed) ──────
    if scored_df is not None and not scored_df.empty:
        summary_df = scored_df[scored_df["score"] >= 2].copy()
        if not summary_df.empty:
            summary_df["metabolite_name"]  = summary_df["metabolite"].map(casic_map)
            summary_df["class"]            = summary_df["metabolite"].map(class_map)
            summary_df["mirror_confirmed"] = summary_df["no_mirror_solo"].astype(bool)

            # FC_produced / FC_consumed depend on which direction triggered
            for col in ("fc_solo", "fc_partner", "fc_return", "direction", "no_mirror_solo"):
                if col not in summary_df.columns:
                    summary_df[col] = np.nan

            summary_df["FC_produced"] = summary_df.apply(
                lambda r: r["fc_solo"] if r["direction"] == "consumed_by_partner"
                else r["fc_partner"], axis=1)
            summary_df["FC_consumed"] = summary_df.apply(
                lambda r: r["fc_partner"] if r["direction"] == "consumed_by_partner"
                else r["fc_return"], axis=1)

            summary_df["_abs_fc_cons"] = summary_df["FC_consumed"].abs()
            summary_df = summary_df.sort_values(
                ["score", "_abs_fc_cons"], ascending=[False, False])

            summary_out = summary_df[[
                "metabolite_name", "class", "chain", "direction",
                "score", "category", "FC_produced", "FC_consumed",
                "mirror_confirmed", "reliable",
            ]].rename(columns={
                "metabolite_name":  "Metabolite_Name",
                "class":            "Class",
                "chain":            "Chain",
                "direction":        "Direction",
                "score":            "Score",
                "category":         "Category",
                "FC_produced":      "FC_produced",
                "FC_consumed":      "FC_consumed",
                "mirror_confirmed": "Mirror_confirmed",
                "reliable":         "Reliable",
            })
            summary_out.to_excel(writer, sheet_name="Summary", index=False)

    writer.close()

    # ── Post-write styling ──────────────────────────────────────────────────
    wb = openpyxl.load_workbook(out)

    # Highlight CV > CV_THRESH in light red on the CV_Table sheet
    if "CV_Table" in wb.sheetnames:
        ws = wb["CV_Table"]
        red_fill = PatternFill("solid", fgColor="FFCDD2")
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                try:
                    if float(cell.value) > CV_THRESH:
                        cell.fill = red_fill
                except (TypeError, ValueError):
                    pass

    # Style Summary header + colour rows by score band
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        header_fill = PatternFill("solid", fgColor="263238")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        score_fills = {
            4: PatternFill("solid", fgColor="C8E6C9"),
            3: PatternFill("solid", fgColor="FFE0B2"),
            2: PatternFill("solid", fgColor="FFF9C4"),
        }
        score_col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "Score":
                score_col_idx = idx
                break
        if score_col_idx:
            for row in ws.iter_rows(min_row=2):
                score_cell = row[score_col_idx - 1]
                try:
                    s = int(score_cell.value)
                    fill = score_fills.get(s)
                    if fill:
                        for cell in row:
                            cell.fill = fill
                except (TypeError, ValueError):
                    pass

    wb.save(out)
    print(f"  Saved: {out.name}")
